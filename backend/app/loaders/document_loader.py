"""
文档加载器（Load 阶段）
支持: PDF / Word(docx) / Excel(xlsx) / Markdown(md) / TXT / HTML

⚠️ PDF 优先 unstructured（按 category 语义分块），失败回退 PyPDFLoader。
"""
from __future__ import annotations

import os
from typing import Optional

from langchain_core.documents import Document


SUPPORTED_EXTS = {
    ".pdf", ".docx", ".doc",
    ".xlsx", ".xls",
    ".md", ".markdown",
    ".txt", ".log", ".csv",
    ".html", ".htm",
}


def get_ext(filename: str) -> str:
    return os.path.splitext(filename)[1].lower()


def load_document(file_path: str, file_name: Optional[str] = None) -> list[Document]:
    """
    统一入口：根据文件扩展名选择 Loader，返回 Document 列表。
    """
    ext = get_ext(file_name or file_path)

    # PDF → PyPDFLoader（依赖 pypdf，稳定）
    if ext == ".pdf":
        return _load_pdf(file_path, file_name)

    # Word → Docx2txtLoader（依赖 python-docx，稳定）
    if ext in {".docx", ".doc"}:
        return _load_word(file_path, file_name)

    # Excel → 按 sheet 拆分，每个 sheet 一个 Document
    if ext in {".xlsx", ".xls"}:
        return _load_excel(file_path, file_name)

    # Markdown → 直接按纯文本读（MD 就是文本）
    if ext in {".md", ".markdown"}:
        return _load_text(file_path, file_name, file_type="md")

    # 纯文本
    if ext in {".txt", ".log", ".csv"}:
        return _load_text(file_path, file_name)

    # HTML → BeautifulSoup 解析
    if ext in {".html", ".htm"}:
        return _load_html(file_path, file_name)

    raise ValueError(f"暂不支持的文件类型: {ext}")


# ==================== 各格式加载实现 ====================

def _load_pdf(file_path: str, name: str) -> list[Document]:
    """
    优先 unstructured 按 category（Title/NarrativeText/Table/ListItem 等）语义切块，
    失败回退 PyPDFLoader。unstructured 分出来的每个 Element 就是一个语义 Document，
    后续 chunker 不再需要再打碎，chunk 数量比按页/按递归切割少 30%~50%，检索更准。
    """
    try:
        from unstructured.partition.pdf import partition_pdf

        elements = partition_pdf(
            filename=file_path,
            strategy="fast",          # fast: 纯 pdfminer 文本；不做 OCR/图片，加载快
            include_page_breaks=False,
            infer_table_structure=False,  # 不做 table HTML，依赖更少
        )

        docs: list[Document] = []
        current_section_title = ""  # 跟随最近的 Title，注入到后续段落 metadata
        for el in elements:
            text = getattr(el, "text", "")
            if not text or not text.strip():
                continue
            category = type(el).__name__   # Title / NarrativeText / Table / ListItem / ...
            page = getattr(el, "metadata", None) and el.metadata.page_number or None

            if category == "Title":
                current_section_title = text.strip()

            content = text
            # 表格保留 category 前缀，方便检索时识别
            if category == "Table":
                content = f"[TABLE]\n{text}"

            meta = {
                "source": name,
                "file_type": "pdf",
                "category": category,
                "section_title": current_section_title,
            }
            if page:
                meta["page"] = page
            docs.append(Document(page_content=content, metadata=meta))

        if docs:
            return docs
        # unstructured 没抽出内容，走回退
    except Exception as _e:
        import logging as _lg
        _lg.getLogger(__name__).info(f"[loader] unstructured PDF 失败，回退 PyPDFLoader: {_e}")

    # 回退：按页切
    from langchain_community.document_loaders import PyPDFLoader
    loader = PyPDFLoader(file_path)
    docs = loader.load()
    for d in docs:
        d.metadata.setdefault("source", name)
        d.metadata.setdefault("file_type", "pdf")
        d.metadata.setdefault("category", "Page")
    return docs


def _load_word(file_path: str, name: str) -> list[Document]:
    from langchain_community.document_loaders import Docx2txtLoader
    loader = Docx2txtLoader(file_path)
    docs = loader.load()
    for d in docs:
        d.metadata.setdefault("source", name)
        d.metadata.setdefault("file_type", "docx")
    return docs


def _load_excel(file_path: str, name: str) -> list[Document]:
    """
    用 openpyxl 读每个 sheet，每个 sheet 生成一个 Document。
    大 Excel 优化：
      1. 按 sheet 拆分 → 每个 sheet 独立分块，避免单 Document 过大
      2. read_only=True + data_only=True → 只读模式，内存占用低
      3. 单 sheet 超过 MAX_SHEET_ROWS 行时自动截断（防 OOM）
    """
    from openpyxl import load_workbook
    MAX_SHEET_ROWS = 5000  # 单个 sheet 最多处理 5000 行，超过截断

    wb = load_workbook(file_path, read_only=True, data_only=True)
    docs: list[Document] = []
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
                row_count += 1
                if row_count >= MAX_SHEET_ROWS:
                    rows.append(f"... (超过 {MAX_SHEET_ROWS} 行已截断)")
                    break
        if rows:
            sheet_content = f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows)
            docs.append(Document(
                page_content=sheet_content,
                metadata={
                    "source": name,
                    "file_type": "xlsx",
                    "sheet_name": sheet_name,
                    "sheet_rows": row_count,
                },
            ))
            total_rows += row_count

    wb.close()

    # 如果只有一个 sheet，直接返回
    if len(docs) <= 1:
        return docs

    # 多个 sheet：额外生成一个"索引" Document，列出所有 sheet 名称和行数
    sheet_index = f"=== Excel 文件结构: {name} ===\n"
    for d in docs:
        sn = d.metadata.get("sheet_name", "unknown")
        sr = d.metadata.get("sheet_rows", 0)
        sheet_index += f"  - Sheet「{sn}」: {sr} 行\n"
    docs.insert(0, Document(
        page_content=sheet_index.strip(),
        metadata={"source": name, "file_type": "xlsx", "is_index": True},
    ))

    return docs


def _load_text(file_path: str, name: str, file_type: str = "txt") -> list[Document]:
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        content = f.read()
    return [Document(page_content=content, metadata={"source": name, "file_type": file_type})]


def _load_html(file_path: str, name: str) -> list[Document]:
    """用 BeautifulSoup 提取纯文本，降级为纯文本读"""
    try:
        from bs4 import BeautifulSoup
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            soup = BeautifulSoup(f, "lxml")
        # 去掉 script/style 标签
        for tag in soup(["script", "style"]):
            tag.decompose()
        text = soup.get_text(separator="\n", strip=True)
        return [Document(page_content=text, metadata={"source": name, "file_type": "html"})]
    except Exception:
        return _load_text(file_path, name, file_type="html")
